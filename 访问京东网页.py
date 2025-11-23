"""
使用 DrissionPage 模块访问京东网页
"""
from DrissionPage import ChromiumPage
import time
import re
from openpyxl import Workbook
from datetime import datetime

# 创建页面对象
page = ChromiumPage()

# 目标网页URL
url = 'https://pro.jd.com/mall/active/2cqCAHDDWm2f5u5irY4TsPPWHQNj/index.html?babelChannel=ttt14&cu=true&utm_source=baidu-pinzhuan&utm_medium=cpc&utm_campaign=t_288551095_baidupinzhuan&utm_term=0f3d30c8dba7459bb52f2eb5eba8ac7d_0_455802cdbb0543e4832e319e90bf9d9e'

try:
    print("正在访问网页...")
    # 访问网页
    page.get(url)
    print(f"成功访问网页: {url}")
    print(f"当前页面标题: {page.title}")
    
    # 等待页面加载完成
    page.wait.load_start()
    
    print("页面加载完成！")
    
    # 等待页面元素加载
    time.sleep(2)
    
    # 查找并点击"京东首页"链接
    print("正在查找京东首页链接...")
    try:
        # 方法1: 通过文本内容查找
        home_link = page.ele('text:京东首页', timeout=5)
        if home_link:
            print("找到京东首页链接，准备点击...")
            home_link.click()
            print("已点击京东首页链接！")
        else:
            # 方法2: 通过aria-label属性查找
            home_link = page.ele('@aria-label:京东首页', timeout=5)
            if home_link:
                print("通过aria-label找到京东首页链接，准备点击...")
                home_link.click()
                print("已点击京东首页链接！")
            else:
                # 方法3: 通过href属性查找
                home_link = page.ele('@href://www.jd.com', timeout=5)
                if home_link:
                    print("通过href找到京东首页链接，准备点击...")
                    home_link.click()
                    print("已点击京东首页链接！")
                else:
                    print("未找到京东首页链接")
    except Exception as e:
        print(f"查找或点击京东首页链接时出错: {e}")
        # 尝试其他定位方式
        try:
            # 尝试通过XPath查找
            home_link = page.ele('xpath://a[@aria-label="京东首页"]', timeout=3)
            if home_link:
                home_link.click()
                print("已通过XPath定位并点击京东首页链接！")
            else:
                # 尝试通过包含jd.com的href查找
                home_link = page.ele('xpath://a[contains(@href, "jd.com")]', timeout=3)
                if home_link:
                    home_link.click()
                    print("已通过href包含jd.com定位并点击链接！")
        except Exception as e2:
            print(f"无法定位京东首页链接: {e2}")
    
    # 等待页面跳转
    time.sleep(2)
    print("等待页面跳转完成...")
    
    # 检查当前窗口数量和URL
    print(f"当前窗口数量: {len(page.tab_ids)}")
    print(f"当前窗口URL: {page.url}")
    
    # 检查是否有新窗口打开，如果有则关闭新窗口，保持在第一个窗口操作
    try:
        print("检查是否有新窗口打开...")
        new_tab = page.wait.new_tab(timeout=5)
        if new_tab:
            print(f"✓ 检测到新窗口")
            print(f"新窗口URL: {new_tab.url}")
            
            # 关闭新窗口，保持在第一个窗口
            print("关闭新窗口，保持在第一个窗口操作...")
            try:
                new_tab.close()
                time.sleep(1)
                print(f"✓ 已关闭新窗口")
            except Exception as close_error:
                print(f"关闭新窗口失败: {close_error}")
        else:
            print("未检测到新窗口，可能在同一窗口跳转")
    except Exception as e:
        print(f"等待新窗口失败: {e}")
        # 如果等待新窗口失败，检查是否有多个窗口
        if len(page.tab_ids) > 1:
            print(f"检测到多个窗口，关闭多余的窗口...")
            try:
                # 获取所有标签页
                tabs = page.tabs
                print(f"共有 {len(tabs)} 个标签页")
                
                # 获取当前窗口的ID
                current_tab_id = None
                try:
                    current_tab_id = page.tab_id
                except:
                    pass
                
                # 关闭除第一个窗口外的所有窗口
                for i, tab in enumerate(tabs):
                    try:
                        tab_id = tab.tab_id if hasattr(tab, 'tab_id') else None
                        tab_url = tab.url
                        print(f"  标签页 {i+1}: URL = {tab_url}, ID = {tab_id}")
                        
                        # 如果不是第一个窗口，关闭它
                        if i > 0:  # 保留第一个窗口（索引0），关闭其他的
                            print(f"  关闭标签页 {i+1}...")
                            tab.close()
                            time.sleep(0.5)
                    except Exception as tab_error:
                        print(f"  处理标签页 {i+1} 时出错: {tab_error}")
                
                print(f"✓ 已关闭多余的窗口，保持在第一个窗口")
                print(f"当前窗口URL: {page.url}")
            except Exception as e2:
                print(f"关闭多余窗口失败: {e2}")
        else:
            print("只有一个窗口，继续操作...")
            time.sleep(1)
            print(f"当前窗口URL: {page.url}")
    
    # 确认是否在京东首页
    current_url = page.url
    print(f"\n当前页面URL: {current_url}")
    
    # 如果还在活动页面，直接访问首页
    if 'active' in current_url:
        print("⚠️ 检测到仍在活动页面，直接访问京东首页...")
        page.get('https://www.jd.com')
        time.sleep(3)
        final_url = page.url
        print(f"访问后的URL: {final_url}")
    elif 'jd.com' in current_url and 'active' not in current_url:
        print("✓ 确认已在京东首页")
    else:
        print(f"⚠️ 当前URL可能不是京东首页: {current_url}")
    
    # 等待页面完全加载
    print("\n等待页面完全加载...")
    try:
        page.wait.load_complete(timeout=10)
    except:
        pass
    time.sleep(2)
    print(f"页面加载完成，当前URL: {page.url}")
    
    # 尝试滚动到顶部，确保搜索框可见
    try:
        page.scroll.to_top()
        time.sleep(1)
    except:
        pass
    
    # 等待页面元素完全加载，特别是等待input元素出现
    print("\n等待页面元素完全加载...")
    try:
        # 等待input#key元素出现
        page.wait.ele_displayed('xpath://input[@id="key"]', timeout=15)
        print("✓ 确认input#key元素已出现在页面上")
        print(f"当前页面URL: {page.url}")
    except Exception as e:
        print(f"⚠️ 等待input#key元素超时: {e}")
        print(f"当前页面URL: {page.url}")
        # 列出所有input元素供调试
        try:
            all_inputs = page.eles('tag:input', timeout=5)
            print(f"页面上共有 {len(all_inputs)} 个input元素")
            for i, inp in enumerate(all_inputs[:10]):
                inp_id = inp.attr('id') or '无id'
                inp_type = inp.attr('type') or '无type'
                print(f"  Input {i+1}: id='{inp_id}', type='{inp_type}'")
        except Exception as e2:
            print(f"无法列出input元素: {e2}")
    
    time.sleep(1)  # 再等待1秒确保元素完全渲染
    print(f"准备查找输入框，当前URL: {page.url}\n")
    
    # 查找搜索输入框并输入内容
    print("正在查找搜索输入框...")
    search_input = None
    try:
        # 方法1: 通过XPath查找input[@id='key']（最准确，确保是input元素）
        print("尝试方法1: 通过XPath查找input[@id='key']...")
        search_input = page.ele('xpath://input[@id="key"]', timeout=10)
        if search_input:
            print(f"  找到元素，标签: {search_input.tag}")
            if search_input.tag.lower() == 'input':
                print("✓ 通过XPath找到正确的input元素")
            else:
                print(f"✗ 找到的元素不是input，而是{search_input.tag}，继续查找...")
                search_input = None
        else:
            print("✗ 未找到元素")
            search_input = None
        
        if not search_input:
            # 方法2: 通过id查找，但验证是input
            print("尝试方法2: 通过id='key'查找（可能找到非input元素）...")
            temp_element = page.ele('#key', timeout=10)
            if temp_element:
                print(f"  找到元素，标签: {temp_element.tag}")
                if temp_element.tag.lower() == 'input':
                    search_input = temp_element
                    print("✓ 通过id找到input元素")
                else:
                    print(f"✗ 通过id找到的是{temp_element.tag}元素，不是input，继续查找...")
                    search_input = None
            else:
                print("✗ 未找到元素")
                search_input = None
        
        if not search_input:
            # 方法3: 通过XPath查找input[@type='text' and @aria-label='搜索']
            print("尝试方法3: 通过XPath查找input[@type='text' and @aria-label='搜索']...")
            search_input = page.ele('xpath://input[@type="text" and @aria-label="搜索"]', timeout=10)
            if search_input and search_input.tag.lower() == 'input':
                print("✓ 通过XPath组合找到input元素")
            else:
                search_input = None
        
        if not search_input:
            # 方法4: 通过aria-label查找，但验证是input
            print("尝试方法4: 通过aria-label='搜索'查找input元素...")
            inputs_with_aria = page.eles('xpath://input[@aria-label="搜索"]', timeout=10)
            if inputs_with_aria:
                for inp in inputs_with_aria:
                    if inp.tag.lower() == 'input':
                        search_input = inp
                        print(f"✓ 通过aria-label找到input元素 (id={inp.attr('id')})")
                        break
                if not search_input:
                    print("✗ 找到的元素都不是input类型")
            else:
                print("✗ 未找到匹配的input元素")
                search_input = None
        
        # 如果还是没找到，列出页面上所有input元素供调试
        if not search_input:
            print("\n" + "=" * 50)
            print("未找到input#key元素，列出页面上所有input元素供参考：")
            print("=" * 50)
            try:
                all_inputs = page.eles('tag:input', timeout=5)
                print(f"页面上共有 {len(all_inputs)} 个input元素：")
                for i, inp in enumerate(all_inputs[:15]):  # 只显示前15个
                    inp_id = inp.attr('id') or '无id'
                    inp_type = inp.attr('type') or '无type'
                    inp_class = inp.attr('class') or '无class'
                    inp_aria = inp.attr('aria-label') or '无aria-label'
                    print(f"  Input {i+1}: id='{inp_id}', type='{inp_type}', class='{inp_class}', aria-label='{inp_aria}'")
            except Exception as e:
                print(f"无法列出input元素: {e}")
            print("=" * 50 + "\n")
        
        if search_input:
            print("=" * 50)
            print("找到元素！")
            print("=" * 50)
            print(f"元素对象: {search_input}")
            print(f"元素类型: {type(search_input)}")
            print(f"元素标签: {search_input.tag}")
            
            # 安全地获取元素属性
            try:
                element_id = search_input.attr('id') or search_input.attrs.get('id', '无')
            except:
                element_id = '无法获取'
            
            try:
                element_classes = search_input.classes if hasattr(search_input, 'classes') else search_input.attrs.get('class', '无')
            except:
                element_classes = '无法获取'
            
            try:
                element_attrs = search_input.attrs
            except:
                element_attrs = '无法获取'
            
            try:
                element_text = search_input.text
            except:
                element_text = '无法获取'
            
            try:
                element_value = search_input.value if hasattr(search_input, 'value') else '无value属性'
            except:
                element_value = '无法获取'
            
            print(f"元素ID: {element_id}")
            print(f"元素class: {element_classes}")
            print(f"元素属性: {element_attrs}")
            print(f"元素文本: {element_text}")
            print(f"元素值: {element_value}")
            print("=" * 50)
            
            # 检查是否是input元素
            if search_input.tag.lower() != 'input':
                print(f"⚠️ 警告：找到的元素不是input标签，而是{search_input.tag}标签！")
                print("继续查找真正的input元素...")
                search_input = None
                
                # 重新查找，只查找input元素
                print("重新查找：只查找input[@id='key']...")
                search_input = page.ele('xpath://input[@id="key"]', timeout=5)
                if search_input and search_input.tag.lower() == 'input':
                    print("✓ 找到正确的input元素！")
                else:
                    # 查找所有input元素
                    print("查找页面上所有input元素...")
                    all_inputs = page.eles('tag:input', timeout=5)
                    print(f"找到 {len(all_inputs)} 个input元素")
                    for inp in all_inputs:
                        inp_id = inp.attr('id') or inp.attrs.get('id', '')
                        inp_type = inp.attr('type') or inp.attrs.get('type', '')
                        inp_aria = inp.attr('aria-label') or inp.attrs.get('aria-label', '')
                        print(f"  - input: id='{inp_id}', type='{inp_type}', aria-label='{inp_aria}'")
                        if inp_id == 'key' or inp_aria == '搜索':
                            search_input = inp
                            print(f"  ✓ 找到匹配的input元素！")
                            break
            
            if search_input and search_input.tag.lower() == 'input':
                print("=" * 50)
                print("找到正确的搜索输入框（input元素）！")
                print("=" * 50)
                print("准备输入内容...")
                
                try:
                    # 等待输入框可见
                    print("等待输入框准备就绪...")
                    time.sleep(0.5)
                    
                    # 滚动到输入框位置，确保可见
                    try:
                        print("滚动到输入框位置...")
                        search_input.scroll.to_see()
                        time.sleep(0.5)
                    except:
                        print("滚动失败，继续尝试...")
                    
                    # 检查输入框当前值
                    try:
                        current_value = search_input.value
                        print(f"输入框当前值: '{current_value}'")
                    except:
                        print("无法获取输入框当前值")
                    
                    # 方法1: 先点击输入框获取焦点，然后输入
                    print("\n尝试方法1: 点击输入框获取焦点，然后使用input()方法...")
                    try:
                        # 点击输入框获取焦点
                        search_input.click()
                        time.sleep(0.5)
                        print("✓ 已点击输入框获取焦点")
                        
                        # 清空输入框
                        search_input.clear()
                        time.sleep(0.3)
                        print("✓ 已清空输入框")
                        
                        # 使用input方法输入
                        search_input.input('RTX 5090D V2显卡')
                        time.sleep(0.8)
                        
                        # 验证输入是否成功
                        input_value = search_input.value
                        print(f"输入后，输入框的值: '{input_value}'")
                        if input_value and ('rtx' in input_value.lower() and '5090' in input_value.lower()):
                            print(f"✓ 已通过input方法输入搜索内容：{input_value}")
                            
                            # 使用JavaScript确保值被设置并触发所有事件
                            print("\n使用JavaScript确保页面显示输入内容...")
                            try:
                                # 方法A: 直接设置值并触发所有事件
                                result = page.run_js('''
                                    var input = document.getElementById('key');
                                    if (input) {
                                        // 先聚焦
                                        input.focus();
                                        
                                        // 设置值
                                        input.value = 'RTX 5090D V2显卡';
                                        
                                        // 触发所有可能的事件
                                        var events = ['input', 'change', 'keyup', 'keydown', 'focus', 'blur'];
                                        events.forEach(function(eventType) {
                                            var event;
                                            if (eventType === 'keyup' || eventType === 'keydown') {
                                                event = new KeyboardEvent(eventType, { 
                                                    bubbles: true, 
                                                    cancelable: true,
                                                    key: 'a',
                                                    code: 'KeyA'
                                                });
                                            } else {
                                                event = new Event(eventType, { 
                                                    bubbles: true, 
                                                    cancelable: true 
                                                });
                                            }
                                            input.dispatchEvent(event);
                                        });
                                        
                                        // 再次聚焦确保可见
                                        input.focus();
                                        
                                        return {
                                            success: true,
                                            value: input.value,
                                            isFocused: document.activeElement === input
                                        };
                                    }
                                    return { success: false };
                                ''')
                                
                                time.sleep(0.8)
                                
                                if result:
                                    if isinstance(result, dict):
                                        if result.get('success'):
                                            print(f"✓ JavaScript设置成功")
                                            print(f"  值: {result.get('value')}")
                                            print(f"  是否聚焦: {result.get('isFocused')}")
                                        else:
                                            print("⚠️ JavaScript设置返回success=false")
                                    else:
                                        print(f"✓ JavaScript执行完成，返回: {result}")
                                else:
                                    print("⚠️ JavaScript设置可能失败")
                                
                                # 再次验证值
                                final_value = search_input.value
                                print(f"最终验证，输入框的值: '{final_value}'")
                                
                                # 等待一下让页面更新
                                time.sleep(0.5)
                                
                                # 使用JavaScript检查页面上实际显示的内容
                                try:
                                    displayed_value = page.run_js('''
                                        var input = document.getElementById('key');
                                        if (input) {
                                            return {
                                                value: input.value,
                                                displayValue: input.value || '',
                                                placeholder: input.placeholder || '',
                                                isVisible: input.offsetWidth > 0 && input.offsetHeight > 0,
                                                style: window.getComputedStyle(input).display
                                            };
                                        }
                                        return null;
                                    ''')
                                    if displayed_value:
                                        print(f"页面上input元素的实际状态:")
                                        if isinstance(displayed_value, dict):
                                            print(f"  value: '{displayed_value.get('value', '')}'")
                                            print(f"  displayValue: '{displayed_value.get('displayValue', '')}'")
                                            print(f"  placeholder: '{displayed_value.get('placeholder', '')}'")
                                            print(f"  isVisible: {displayed_value.get('isVisible', False)}")
                                            print(f"  display style: '{displayed_value.get('style', '')}'")
                                        else:
                                            print(f"  {displayed_value}")
                                except Exception as check_error:
                                    print(f"检查页面显示失败: {check_error}")
                                    
                            except Exception as js_error:
                                print(f"JavaScript触发事件失败: {js_error}")
                                import traceback
                                traceback.print_exc()
                        else:
                            print(f"✗ input方法失败，输入框值仍为: '{input_value}'")
                            raise Exception("input方法未成功")
                    except Exception as method1_error:
                        print(f"方法1失败: {method1_error}")
                        
                        # 方法2: 使用set.value方法
                        print("\n尝试方法2: 使用set.value()方法...")
                        try:
                            search_input.set.value('RTX 5090D V2显卡')
                            time.sleep(0.8)
                            
                            # 验证输入是否成功
                            input_value = search_input.value
                            print(f"输入后，输入框的值: '{input_value}'")
                            if input_value and ('rtx' in input_value.lower() and '5090' in input_value.lower()):
                                print(f"✓ 已通过set.value输入搜索内容：{input_value}")
                            else:
                                print(f"✗ set.value方法失败，输入框值仍为: '{input_value}'")
                                raise Exception("set.value方法未成功")
                        except Exception as method2_error:
                            print(f"方法2失败: {method2_error}")
                            
                            # 方法3: 先点击，再清空，再逐个字符输入
                            print("\n尝试方法3: 逐个字符输入...")
                            try:
                                # 点击获取焦点
                                search_input.click()
                                time.sleep(0.3)
                                
                                # 清空
                                search_input.clear()
                                time.sleep(0.3)
                                
                                # 逐个字符输入
                                text_to_input = 'RTX 5090D V2显卡'
                                for char in text_to_input:
                                    search_input.input(char)
                                    time.sleep(0.1)
                                
                                time.sleep(0.5)
                                
                                # 验证输入是否成功
                                input_value = search_input.value
                                print(f"输入后，输入框的值: '{input_value}'")
                                if input_value and ('rtx' in input_value.lower() and '5090' in input_value.lower()):
                                    print(f"✓ 已通过逐个字符输入搜索内容：{input_value}")
                                else:
                                    print(f"✗ 逐个字符输入也失败，输入框值仍为: '{input_value}'")
                            except Exception as method3_error:
                                print(f"方法3失败: {method3_error}")
                                import traceback
                                traceback.print_exc()
                    
                    # 查找并点击搜索按钮
                    print("正在查找搜索按钮...")
                    time.sleep(0.5)
                
                    search_button = None
                    # 方法1: 通过aria-label查找搜索按钮（确保是button标签）
                    buttons = page.eles('@aria-label:搜索', timeout=5)
                    for btn in buttons:
                        if btn.tag == 'button':
                            search_button = btn
                            break
                    
                    if not search_button:
                        # 方法2: 通过XPath查找button标签
                        search_button = page.ele('xpath://button[@aria-label="搜索"]', timeout=5)
                    
                    if not search_button:
                        # 方法3: 通过class和文本内容查找
                        search_button = page.ele('xpath://button[@class="button" and text()="搜索"]', timeout=5)
                    
                    if not search_button:
                        # 方法4: 通过文本内容查找（可能是第一个button）
                        buttons = page.eles('text:搜索', timeout=5)
                        for btn in buttons:
                            if btn.tag == 'button':
                                search_button = btn
                                break
                    
                    if search_button:
                        # 点击搜索按钮
                        try:
                            search_button.click()
                            print("已点击搜索按钮，开始搜索！")
                        except Exception as click_error:
                            print(f"点击按钮失败: {click_error}")
                            print("请检查按钮是否可见和可点击")
                    else:
                        print("未找到搜索按钮")
                        
                except Exception as e:
                    print(f"操作搜索输入框或按钮时出错: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print("✗ 未找到正确的input输入框元素，无法继续输入")
                search_input = None
        else:
            print("=" * 50)
            print("未找到搜索输入框！")
            print("=" * 50)
            print("已尝试以下所有方法：")
            print("1. 通过id='key'查找")
            print("2. 通过aria-label='搜索'查找")
            print("3. 通过class='text'查找")
            print("4. 通过XPath查找input[@id='key']")
            print("5. 通过XPath查找input[@type='text' and @aria-label='搜索']")
            print("=" * 50)
            # 尝试查找所有input元素，看看页面上有哪些input
            print("尝试查找页面上所有的input元素...")
            all_inputs = page.eles('tag:input', timeout=3)
            print(f"页面上共有 {len(all_inputs)} 个input元素")
            for i, inp in enumerate(all_inputs[:10]):  # 只显示前10个
                print(f"  Input {i+1}: id={inp.id}, type={inp.attrs.get('type', '')}, class={inp.classes}, aria-label={inp.attrs.get('aria-label', '')}")
            print("=" * 50)
    except Exception as e:
        print(f"查找或操作搜索输入框时出错: {e}")
        import traceback
        traceback.print_exc()
    
    # 等待搜索结果加载
    print("\n等待搜索结果页面加载...")
    time.sleep(3)
    print(f"搜索结果页面URL: {page.url}")
    
    # 查找并点击"显卡"分类
    print("\n正在查找'显卡'分类元素...")
    try:
        # 方法1: 通过class和文本内容查找
        graphics_card_span = page.ele('xpath://span[@class="_value-label_1xq81_50" and text()="显卡"]', timeout=10)
        if not graphics_card_span:
            # 方法2: 通过class查找，然后筛选文本内容
            all_spans = page.eles('xpath://span[@class="_value-label_1xq81_50"]', timeout=10)
            for span in all_spans:
                if span.text == '显卡':
                    graphics_card_span = span
                    break
        
        if graphics_card_span:
            print("✓ 找到'显卡'分类元素")
            print(f"元素文本: {graphics_card_span.text}")
            
            # 滚动到元素位置
            try:
                graphics_card_span.scroll.to_see()
                time.sleep(0.5)
            except:
                pass
            
            # 点击"显卡"
            graphics_card_span.click()
            print("✓ 已点击'显卡'分类")
            time.sleep(2)  # 等待页面更新
        else:
            print("✗ 未找到'显卡'分类元素")
    except Exception as e:
        print(f"查找或点击'显卡'分类时出错: {e}")
        import traceback
        traceback.print_exc()
    
    # 查找并双击"价格"排序元素
    print("\n正在查找'价格'排序元素...")
    try:
        price_element = None
        
        # 方法1: 直接查找包含"价格"文本的span元素
        print("方法1: 查找span[text()='价格']...")
        price_span = page.ele('xpath://span[text()="价格"]', timeout=10)
        if price_span:
            print(f"✓ 找到span元素，文本: {price_span.text}")
            # 找到span的父元素div（class="_sort-tag-inner_3m6t1_24"）
            try:
                price_element = price_span.parent()
                if price_element and '_sort-tag-inner_3m6t1_24' in price_element.attr('class'):
                    print(f"✓ 找到父元素div")
                else:
                    # 如果父元素不对，尝试向上查找
                    price_element = price_span
                    print("使用span元素本身")
            except:
                price_element = price_span
        
        if not price_element:
            # 方法2: 查找所有包含"价格"文本的span，然后找到对应的div
            print("方法2: 查找所有包含'价格'的span元素...")
            all_spans = page.eles('xpath://span[contains(text(), "价格")]', timeout=10)
            for span in all_spans:
                if span.text == '价格':
                    print(f"找到span，文本: {span.text}")
                    try:
                        parent = span.parent()
                        if parent and '_sort-tag-inner_3m6t1_24' in parent.attr('class'):
                            price_element = parent
                            print(f"✓ 找到父元素div")
                            break
                        else:
                            price_element = span
                            print("使用span元素本身")
                            break
                    except:
                        price_element = span
                        break
        
        if not price_element:
            # 方法3: 查找div中包含"价格"span的
            print("方法3: 查找div中包含'价格'span的元素...")
            all_divs = page.eles('xpath://div[@class="_sort-tag-inner_3m6t1_24"]', timeout=10)
            for div in all_divs:
                # 检查div中是否包含"价格"文本的span
                try:
                    spans = div.eles('tag:span')
                    for span in spans:
                        if span.text == '价格':
                            price_element = div
                            print(f"✓ 找到包含'价格'span的div")
                            break
                    if price_element:
                        break
                except:
                    if '价格' in div.text:
                        price_element = div
                        print(f"✓ 找到包含'价格'文本的div")
                        break
        
        if price_element:
            print(f"✓ 找到'价格'排序元素")
            print(f"元素标签: {price_element.tag}")
            print(f"元素文本: {price_element.text}")
            
            # 滚动到元素位置
            try:
                price_element.scroll.to_see()
                time.sleep(0.5)
            except:
                pass
            
            # 双击"价格"（点击两次）
            print("准备双击'价格'排序...")
            
            # 第一次点击
            print("第一次点击'价格'...")
            try:
                price_element.click()
                print("✓ 第一次点击成功")
                time.sleep(1)  # 等待第一次点击生效
            except Exception as click1_error:
                print(f"✗ 第一次点击失败: {click1_error}")
            
            # 重新查找元素，确保元素仍然可用（页面可能已更新）
            try:
                # 重新查找价格元素
                price_span = page.ele('xpath://span[text()="价格"]', timeout=5)
                if price_span:
                    try:
                        price_element = price_span.parent()
                        if price_element and '_sort-tag-inner_3m6t1_24' in price_element.attr('class'):
                            pass  # 找到了
                        else:
                            price_element = price_span
                    except:
                        price_element = price_span
                else:
                    # 如果找不到，尝试其他方法
                    all_divs = page.eles('xpath://div[@class="_sort-tag-inner_3m6t1_24"]', timeout=5)
                    for div in all_divs:
                        spans = div.eles('tag:span')
                        for span in spans:
                            if span.text == '价格':
                                price_element = div
                                break
                        if price_element:
                            break
            except:
                print("重新查找元素失败，使用原元素继续...")
            
            # 第二次点击
            print("第二次点击'价格'...")
            try:
                price_element.click()
                print("✓ 第二次点击成功")
                time.sleep(1)  # 等待第二次点击生效
            except Exception as click2_error:
                print(f"✗ 第二次点击失败: {click2_error}")
            
            print("✓ 已双击'价格'排序")
            time.sleep(2)  # 等待页面更新
        else:
            print("✗ 未找到'价格'排序元素")
            # 列出所有可能的元素供调试
            try:
                print("列出所有包含'价格'的元素...")
                all_price_elements = page.eles('xpath://*[contains(text(), "价格")]', timeout=5)
                for i, elem in enumerate(all_price_elements[:10]):
                    print(f"  元素 {i+1}: 标签={elem.tag}, 文本={elem.text}, class={elem.attr('class')}")
            except:
                pass
    except Exception as e:
        print(f"查找或点击'价格'排序时出错: {e}")
        import traceback
        traceback.print_exc()
    
    # 获取前10个商品名称并保存到Excel
    print("\n开始获取商品信息...")
    try:
        # 等待页面完全加载
        print("等待商品列表加载...")
        time.sleep(3)
        
        # 查找所有商品标签
        print("查找商品标签...")
        product_wrappers = page.eles('xpath://div[@class="_wrapper_8v3rv_3 plugin_goodsCardWrapper _row_6_8v3rv_13"]', timeout=10)
        
        if not product_wrappers:
            # 尝试其他可能的class组合
            product_wrappers = page.eles('xpath://div[contains(@class, "_wrapper_8v3rv_3") and contains(@class, "plugin_goodsCardWrapper")]', timeout=10)
        
        print(f"找到 {len(product_wrappers)} 个商品")
        
        # 获取所有商品的名称、价格、显卡型号、显存大小和店铺名称
        product_data = []  # 存储商品信息：[(名称, 显卡型号, 显存大小, 价格, 店铺名称), ...]
        for i, wrapper in enumerate(product_wrappers):
            product_name = ''
            product_price = ''
            gpu_model = ''
            memory_size = ''
            shop_name = ''
            
            try:
                # 在每个商品标签下查找商品名称
                name_span = wrapper.ele('xpath:.//span[@class="_text_1g56m_31"]', timeout=3)
                if name_span:
                    # 获取商品名称（去除HTML标签，只保留文本）
                    product_name = name_span.text
                    if not product_name:
                        # 如果text为空，尝试获取title属性
                        product_name = name_span.attr('title') or ''
                    
                    # 从商品名称中提取显卡型号和显存大小
                    if product_name:
                        # 提取显卡型号
                        model_patterns = [
                            r'RTX\s*5090[Dd]\s*[Vv]?\s*2',  # RTX 5090D V2, RTX5090D v2
                            r'RTX\s*5090[Dd]',  # RTX 5090D
                            r'5090[Dd]\s*[Vv]?\s*2',  # 5090D V2, 5090D v2
                        ]
                        for pattern in model_patterns:
                            match = re.search(pattern, product_name, re.IGNORECASE)
                            if match:
                                gpu_model = match.group().strip()
                                break
                        
                        # 如果没找到，尝试查找包含5090的型号
                        if not gpu_model:
                            match = re.search(r'RTX\s*\d+[Dd]?', product_name, re.IGNORECASE)
                            if match:
                                gpu_model = match.group().strip()
                        
                        # 提取显存大小（如：24G、32G、24GB、32GB等）
                        memory_patterns = [
                            r'(\d+)\s*[Gg][Bb]?',  # 24G, 32GB, 24GB等
                            r'(\d+)\s*[Gg]',  # 24G, 32G等
                        ]
                        for pattern in memory_patterns:
                            match = re.search(pattern, product_name)
                            if match:
                                memory_size = match.group(1) + 'G'  # 统一格式为数字+G
                                break
                
                # 查找价格元素
                price_span = wrapper.ele('xpath:.//span[@class="_price_uqsva_14"]', timeout=3)
                if price_span:
                    # 获取价格文本（包含¥符号和数字）
                    price_text = price_span.text
                    if price_text:
                        # 提取数字部分（去除¥符号和空格）
                        price_match = re.search(r'[\d,]+', price_text.replace(',', ''))
                        if price_match:
                            product_price = price_match.group()
                        else:
                            product_price = price_text.strip()
                    else:
                        # 如果text为空，尝试获取内部文本
                        try:
                            # 查找所有子元素，提取价格数字
                            price_elements = price_span.eles('tag:*')
                            for elem in price_elements:
                                elem_text = elem.text or ''
                                price_match = re.search(r'[\d,]+', elem_text.replace(',', ''))
                                if price_match:
                                    product_price = price_match.group()
                                    break
                        except:
                            pass
                
                # 查找店铺名称
                shop_link = wrapper.ele('xpath:.//a[@class="_name_d19t5_35"]', timeout=3)
                if shop_link:
                    # 在店铺链接下查找店铺名称span
                    shop_span = shop_link.ele('xpath:.//span', timeout=2)
                    if shop_span:
                        shop_name = shop_span.text or ''
                
                if product_name:
                    product_data.append((product_name, gpu_model, memory_size, product_price, shop_name))
                    print(f"  商品 {i+1}: {product_name}")
                    print(f"    显卡型号: {gpu_model if gpu_model else '未提取到'}")
                    print(f"    显存大小: {memory_size if memory_size else '未提取到'}")
                    print(f"    今日最低价: {product_price if product_price else '未获取到'}")
                    print(f"    店铺名称: {shop_name if shop_name else '未获取到'}")
                else:
                    print(f"  商品 {i+1}: 无法获取名称")
            except Exception as e:
                print(f"  商品 {i+1}: 获取信息时出错 - {e}")
                import traceback
                traceback.print_exc()
        
        # 找到价格最低的商品
        if product_data:
            print(f"\n共获取到 {len(product_data)} 个商品信息，正在查找价格最低的商品...")
            
            # 筛选出有价格的商品并找到最低价
            valid_products = []
            for name, model, memory, price, shop in product_data:
                if price and price.isdigit():
                    try:
                        price_num = int(price)
                        valid_products.append((name, model, memory, price, shop, price_num))
                    except:
                        pass
            
            if valid_products:
                # 按价格排序，找到最低价的商品
                valid_products.sort(key=lambda x: x[5])  # 按价格排序
                lowest_price_product = valid_products[0]
                
                print(f"\n找到价格最低的商品:")
                print(f"  商品名称: {lowest_price_product[0]}")
                print(f"  显卡型号: {lowest_price_product[1] if lowest_price_product[1] else '未提取到'}")
                print(f"  显存大小: {lowest_price_product[2] if lowest_price_product[2] else '未提取到'}")
                print(f"  今日最低价: {lowest_price_product[3]}")
                print(f"  店铺名称: {lowest_price_product[4] if lowest_price_product[4] else '未获取到'}")
                
                # 创建Excel工作簿
                wb = Workbook()
                ws = wb.active
                ws.title = "商品列表"
                
                # 获取当前日期时间
                save_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                platform = "京东"
                
                # 设置表头（调整顺序：电商平台在店铺名称前）
                ws['A1'] = '序号'
                ws['B1'] = '商品名称'
                ws['C1'] = '显卡型号'
                ws['D1'] = '显存大小'
                ws['E1'] = '今日最低价'
                ws['F1'] = '保存日期'
                ws['G1'] = '电商平台'
                ws['H1'] = '店铺名称'
                
                # 写入数据（只保存价格最低的一个商品）
                name, model, memory, price, shop = lowest_price_product[:5]
                ws['A2'] = 1
                ws['B2'] = name
                ws['C2'] = model if model else '未提取到'
                ws['D2'] = memory if memory else '未提取到'
                ws['E2'] = price if price else '未获取到'
                ws['F2'] = save_date
                ws['G2'] = platform
                ws['H2'] = shop if shop else '未获取到'
                
                # 调整列宽
                ws.column_dimensions['A'].width = 10
                ws.column_dimensions['B'].width = 100
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 20
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 30
                
                # 生成文件名（包含时间戳）
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"商品列表_{timestamp}.xlsx"
                
                # 保存文件
                wb.save(filename)
                print(f"\n✓ 已保存到文件: {filename}")
                print(f"  保存了价格最低的1条商品信息")
                print(f"  保存日期: {save_date}")
                print(f"  电商平台: {platform}")
            else:
                print("✗ 未找到有有效价格的商品")
        else:
            print("✗ 未获取到任何商品信息")
            
    except Exception as e:
        print(f"获取商品信息时出错: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n操作完成！")
    
    # 保持浏览器打开，方便查看
    input("按 Enter 键关闭浏览器...")
    
except Exception as e:
    print(f"访问网页时出错: {e}")
    
finally:
    # 关闭页面
    page.close()
    print("浏览器已关闭")

